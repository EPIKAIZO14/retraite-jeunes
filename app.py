from flask import Flask, render_template, request, redirect, url_for, session, send_file
from supabase import create_client, Client
import openpyxl
import io

app = Flask(__name__)
app.secret_key = 'cle_secrete_super_securisee'

SUPABASE_URL = "https://solmrldhvdqnsvpnqhxg.supabase.co"
SUPABASE_KEY = "sb_publishable_zfePN75CDKoVXkq67Mhxvw_KiECA9KC"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Mot de passe pour accéder à la liste sécurisée
ADMIN_PASSWORD = "admin"

@app.route('/', methods=['GET', 'POST'])
def index():
    success = False
    if request.method == 'POST':
        data = {
            "nom": request.form.get('nom'),
            "post_nom": request.form.get('post_nom'),
            "prenom": request.form.get('prenom'),
            "genre": request.form.get('genre'),
            "telephone": request.form.get('telephone'),
            "adresse": request.form.get('adresse')
        }
        supabase.table("gestion").insert(data).execute()
        success = True
    return render_template('index.html', success=success)

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('liste'))
        else:
            error = "Mot de passe incorrect !"
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

@app.route('/liste')
def liste():
    if not session.get('admin'):
        return redirect(url_for('login'))
    
    response = supabase.table("gestion").select("*").execute()
    rows = response.data if response.data else []
    return render_template('liste.html', rows=rows)

@app.route('/download-excel')
def download_excel():
    if not session.get('admin'):
        return redirect(url_for('login'))
    
    response = supabase.table("gestion").select("*").execute()
    rows = response.data if response.data else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jeunes Inscrits"

    headers = ["N°", "Nom", "Post-nom", "Prénom", "Genre", "Téléphone", "Adresse"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    for index, row in enumerate(rows, start=1):
        ws.append([
            index,
            row.get("nom"),
            row.get("post_nom"),
            row.get("prenom"),
            row.get("genre"),
            row.get("telephone"),
            row.get("adresse")
        ])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='liste_jeunes_inscrits.xlsx'
    )

@app.route('/supprimer/<int:id>')
def supprimer(id):
    if not session.get('admin'):
        return redirect(url_for('login'))
    supabase.table("gestion").delete().eq("id", id).execute()
    return redirect(url_for('liste'))

if __name__ == '__main__':
    app.run(debug=True)